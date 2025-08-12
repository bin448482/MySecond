"""
增强版输出管理器
优化报告格式，集成技术指标、评分系统和策略引擎的输出
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
import logging
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from database import DatabaseManager
from enhanced_technical_indicators import EnhancedTechnicalIndicators
from comprehensive_scoring import ComprehensiveScoring
from stock_strategy_engine import StockStrategyEngine
from strategy_backtest import StrategyBacktest
from utils import config_manager

logger = logging.getLogger(__name__)


class EnhancedOutputManager:
    """增强版输出管理器"""
    
    def __init__(self, db_manager: DatabaseManager):
        """
        初始化增强版输出管理器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db = db_manager
        self.tech_indicators = EnhancedTechnicalIndicators(db_manager)
        self.scoring = ComprehensiveScoring(db_manager)
        self.strategy_engine = StockStrategyEngine(db_manager)
        self.backtest = StrategyBacktest(db_manager)
        self.config = config_manager
        
        # 确保输出目录存在
        self.output_dir = "output"
        self._ensure_output_dir()
        
        # 样式配置
        self.styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subheader': {
                'font': Font(bold=True, color='000000'),
                'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'positive': {
                'font': Font(color='006100'),
                'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            },
            'negative': {
                'font': Font(color='9C0006'),
                'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            },
            'neutral': {
                'font': Font(color='000000'),
                'alignment': Alignment(horizontal='center')
            }
        }
    
    def _ensure_output_dir(self):
        """确保输出目录存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def _apply_cell_style(self, cell, style_name: str):
        """应用单元格样式"""
        if style_name in self.styles:
            style = self.styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
    
    def _format_percentage(self, value: float, decimal_places: int = 2) -> str:
        """格式化百分比"""
        if pd.isna(value):
            return "N/A"
        return f"{value:.{decimal_places}f}%"
    
    def _format_number(self, value: float, decimal_places: int = 2) -> str:
        """格式化数字"""
        if pd.isna(value):
            return "N/A"
        return f"{value:.{decimal_places}f}"
    
    def create_strategy_selection_report(self, strategy_name: str, 
                                       max_results: int = 50) -> str:
        """
        创建策略选股报告
        
        Args:
            strategy_name: 策略名称
            max_results: 最大结果数
            
        Returns:
            报告文件路径
        """
        try:
            logger.info(f"开始生成策略选股报告: {strategy_name}")
            
            # 执行策略选股
            selection_results = self.strategy_engine.execute_strategy(strategy_name, max_results)
            
            if selection_results.empty:
                logger.warning(f"策略 {strategy_name} 没有选出任何股票")
                return ""
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 1. 创建选股结果工作表
            self._create_selection_results_sheet(wb, selection_results, strategy_name)
            
            # 2. 创建技术指标详情工作表
            self._create_technical_details_sheet(wb, selection_results.head(20))
            
            # 3. 创建策略配置工作表
            self._create_strategy_config_sheet(wb, strategy_name)
            
            # 4. 创建市场概况工作表
            self._create_market_overview_sheet(wb)
            
            # 保存文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{self.output_dir}/strategy_selection_{strategy_name}_{timestamp}.xlsx"
            wb.save(filename)
            
            logger.info(f"策略选股报告已生成: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"生成策略选股报告失败: {e}")
            return ""
    
    def _create_selection_results_sheet(self, wb: Workbook, results: pd.DataFrame, strategy_name: str):
        """创建选股结果工作表"""
        ws = wb.create_sheet(title="选股结果")
        
        # 标题
        ws['A1'] = f"策略选股结果 - {strategy_name}"
        ws.merge_cells('A1:M1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 基本信息
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"选股数量: {len(results)} 只"
        ws['A4'] = f"策略名称: {strategy_name}"
        
        # 表头
        headers = [
            '排名', '股票代码', '股票名称', '当前价格', '综合评分', 
            '技术评分', '动量评分', '成交量评分', '波动率评分',
            '5日涨幅%', '成交量', '换手率%', '选股日期'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            self._apply_cell_style(cell, 'subheader')
        
        # 数据行
        for row_idx, (_, row) in enumerate(results.iterrows(), 7):
            ws.cell(row=row_idx, column=1, value=row_idx - 6)  # 排名
            ws.cell(row=row_idx, column=2, value=row.get('symbol', ''))
            ws.cell(row=row_idx, column=3, value=row.get('name', ''))
            ws.cell(row=row_idx, column=4, value=row.get('close', 0))
            ws.cell(row=row_idx, column=5, value=row.get('comprehensive_score', 0))
            ws.cell(row=row_idx, column=6, value=row.get('technical_score', 0))
            ws.cell(row=row_idx, column=7, value=row.get('momentum_score', 0))
            ws.cell(row=row_idx, column=8, value=row.get('volume_score', 0))
            ws.cell(row=row_idx, column=9, value=row.get('volatility_score', 0))
            ws.cell(row=row_idx, column=10, value=row.get('price_change_5d', 0))
            ws.cell(row=row_idx, column=11, value=row.get('volume', 0))
            ws.cell(row=row_idx, column=12, value=row.get('turnover_rate', 0))
            ws.cell(row=row_idx, column=13, value=row.get('selection_date', ''))
            
            # 应用条件格式
            score_cell = ws.cell(row=row_idx, column=5)
            if row.get('comprehensive_score', 0) >= 80:
                self._apply_cell_style(score_cell, 'positive')
            elif row.get('comprehensive_score', 0) >= 60:
                self._apply_cell_style(score_cell, 'neutral')
            else:
                self._apply_cell_style(score_cell, 'negative')
            
            # 涨幅颜色
            change_cell = ws.cell(row=row_idx, column=10)
            if row.get('price_change_5d', 0) > 0:
                self._apply_cell_style(change_cell, 'positive')
            elif row.get('price_change_5d', 0) < 0:
                self._apply_cell_style(change_cell, 'negative')
        
        # 调整列宽
        column_widths = [6, 12, 15, 10, 10, 10, 10, 10, 10, 10, 12, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    def _create_technical_details_sheet(self, wb: Workbook, top_stocks: pd.DataFrame):
        """创建技术指标详情工作表"""
        ws = wb.create_sheet(title="技术指标详情")
        
        # 标题
        ws['A1'] = "技术指标详情 (前20只股票)"
        ws.merge_cells('A1:P1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 表头
        headers = [
            '股票代码', '股票名称', 'MACD', 'MACD信号', 'RSI', 
            'MA5', 'MA20', 'MA60', '布林上轨', '布林下轨', 
            'KDJ_K', 'KDJ_D', 'CCI', '威廉%R', 'ATR', '量比'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, 'subheader')
        
        # 获取技术指标数据
        for row_idx, (_, stock) in enumerate(top_stocks.iterrows(), 4):
            symbol = stock.get('symbol', '')
            
            try:
                # 获取技术指标
                indicators = self.tech_indicators.calculate_all_indicators(symbol)
                
                if not indicators.empty:
                    latest = indicators.iloc[-1]
                    
                    ws.cell(row=row_idx, column=1, value=symbol)
                    ws.cell(row=row_idx, column=2, value=stock.get('name', ''))
                    ws.cell(row=row_idx, column=3, value=self._format_number(latest.get('macd', 0), 4))
                    ws.cell(row=row_idx, column=4, value=self._format_number(latest.get('macd_signal', 0), 4))
                    ws.cell(row=row_idx, column=5, value=self._format_number(latest.get('rsi', 0), 2))
                    ws.cell(row=row_idx, column=6, value=self._format_number(latest.get('ma5', 0), 2))
                    ws.cell(row=row_idx, column=7, value=self._format_number(latest.get('ma20', 0), 2))
                    ws.cell(row=row_idx, column=8, value=self._format_number(latest.get('ma60', 0), 2))
                    ws.cell(row=row_idx, column=9, value=self._format_number(latest.get('bb_upper', 0), 2))
                    ws.cell(row=row_idx, column=10, value=self._format_number(latest.get('bb_lower', 0), 2))
                    ws.cell(row=row_idx, column=11, value=self._format_number(latest.get('kdj_k', 0), 2))
                    ws.cell(row=row_idx, column=12, value=self._format_number(latest.get('kdj_d', 0), 2))
                    ws.cell(row=row_idx, column=13, value=self._format_number(latest.get('cci', 0), 2))
                    ws.cell(row=row_idx, column=14, value=self._format_number(latest.get('williams_r', 0), 2))
                    ws.cell(row=row_idx, column=15, value=self._format_number(latest.get('atr', 0), 2))
                    ws.cell(row=row_idx, column=16, value=self._format_number(latest.get('volume_ratio', 0), 2))
                    
            except Exception as e:
                logger.error(f"获取股票 {symbol} 技术指标失败: {e}")
                ws.cell(row=row_idx, column=1, value=symbol)
                ws.cell(row=row_idx, column=2, value="数据获取失败")
        
        # 调整列宽
        for col in range(1, 17):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
    
    def _create_strategy_config_sheet(self, wb: Workbook, strategy_name: str):
        """创建策略配置工作表"""
        ws = wb.create_sheet(title="策略配置")
        
        # 标题
        ws['A1'] = f"策略配置 - {strategy_name}"
        ws.merge_cells('A1:D1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 获取策略配置
        strategies = self.strategy_engine.get_available_strategies()
        strategy_config = strategies.get(strategy_name, {})
        
        row = 3
        
        # 基本信息
        ws.cell(row=row, column=1, value="策略名称").font = Font(bold=True)
        ws.cell(row=row, column=2, value=strategy_config.get('name', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="策略描述").font = Font(bold=True)
        ws.cell(row=row, column=2, value=strategy_config.get('description', ''))
        row += 2
        
        # 权重配置
        ws.cell(row=row, column=1, value="权重配置").font = Font(bold=True)
        row += 1
        
        weights = strategy_config.get('weights', {})
        for weight_name, weight_value in weights.items():
            ws.cell(row=row, column=2, value=weight_name)
            ws.cell(row=row, column=3, value=f"{weight_value:.1%}")
            row += 1
        
        row += 1
        
        # 筛选条件
        ws.cell(row=row, column=1, value="筛选条件").font = Font(bold=True)
        row += 1
        
        filters = strategy_config.get('filters', {})
        for filter_name, filter_value in filters.items():
            ws.cell(row=row, column=2, value=filter_name)
            ws.cell(row=row, column=3, value=str(filter_value))
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_market_overview_sheet(self, wb: Workbook):
        """创建市场概况工作表"""
        ws = wb.create_sheet(title="市场概况")
        
        # 标题
        ws['A1'] = "市场概况"
        ws.merge_cells('A1:D1')
        self._apply_cell_style(ws['A1'], 'header')
        
        try:
            # 获取数据库统计信息
            stats = self.db.get_database_stats()
            
            row = 3
            ws.cell(row=row, column=1, value="数据库统计").font = Font(bold=True)
            row += 1
            
            for stat_name, stat_value in stats.items():
                ws.cell(row=row, column=2, value=stat_name)
                ws.cell(row=row, column=3, value=stat_value)
                row += 1
            
            # 获取股票市场分布
            market_stats = self.db.get_stock_count_by_market()
            
            row += 1
            ws.cell(row=row, column=1, value="市场分布").font = Font(bold=True)
            row += 1
            
            for market, count in market_stats.items():
                ws.cell(row=row, column=2, value=market)
                ws.cell(row=row, column=3, value=count)
                row += 1
            
        except Exception as e:
            logger.error(f"获取市场概况数据失败: {e}")
            ws.cell(row=3, column=1, value="数据获取失败")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def create_backtest_report(self, strategy_name: str, 
                             start_date: str, 
                             end_date: str) -> str:
        """
        创建回测报告
        
        Args:
            strategy_name: 策略名称
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            报告文件路径
        """
        try:
            logger.info(f"开始生成回测报告: {strategy_name}")
            
            # 执行回测
            backtest_result = self.backtest.backtest_strategy(strategy_name, start_date, end_date)
            
            if 'error' in backtest_result:
                logger.error(f"回测失败: {backtest_result['error']}")
                return ""
            
            # 创建Excel工作簿
            wb = Workbook()
            wb.remove(wb.active)
            
            # 1. 创建回测概况工作表
            self._create_backtest_overview_sheet(wb, backtest_result)
            
            # 2. 创建表现分析工作表
            self._create_performance_analysis_sheet(wb, backtest_result)
            
            # 3. 创建选股样本工作表
            self._create_selection_samples_sheet(wb, backtest_result)
            
            # 保存文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{self.output_dir}/backtest_report_{strategy_name}_{timestamp}.xlsx"
            wb.save(filename)
            
            logger.info(f"回测报告已生成: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"生成回测报告失败: {e}")
            return ""
    
    def _create_backtest_overview_sheet(self, wb: Workbook, backtest_result: Dict[str, Any]):
        """创建回测概况工作表"""
        ws = wb.create_sheet(title="回测概况")
        
        # 标题
        ws['A1'] = f"回测概况 - {backtest_result.get('strategy', 'Unknown')}"
        ws.merge_cells('A1:D1')
        self._apply_cell_style(ws['A1'], 'header')
        
        row = 3
        
        # 基本信息
        ws.cell(row=row, column=1, value="回测期间").font = Font(bold=True)
        ws.cell(row=row, column=2, value=backtest_result.get('backtest_period', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="测试次数").font = Font(bold=True)
        ws.cell(row=row, column=2, value=backtest_result.get('test_dates', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="总选股数").font = Font(bold=True)
        ws.cell(row=row, column=2, value=backtest_result.get('total_selections', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="回测时间").font = Font(bold=True)
        ws.cell(row=row, column=2, value=backtest_result.get('backtest_date', ''))
        row += 2
        
        # 表现总结
        performance = backtest_result.get('performance', {})
        summary = performance.get('summary', {})
        
        if summary:
            ws.cell(row=row, column=1, value="表现总结").font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=2, value="主要收益率")
            ws.cell(row=row, column=3, value=self._format_percentage(summary.get('primary_avg_return', 0)))
            row += 1
            
            ws.cell(row=row, column=2, value="胜率")
            ws.cell(row=row, column=3, value=self._format_percentage(summary.get('primary_positive_rate', 0)))
            row += 1
            
            ws.cell(row=row, column=2, value="夏普比率")
            ws.cell(row=row, column=3, value=self._format_number(summary.get('sharpe_ratio', 0)))
            row += 1
            
            ws.cell(row=row, column=2, value="策略评级")
            ws.cell(row=row, column=3, value=summary.get('strategy_rating', 'N/A'))
            row += 1
            
            ws.cell(row=row, column=2, value="风险水平")
            ws.cell(row=row, column=3, value=summary.get('risk_level', 'N/A'))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_performance_analysis_sheet(self, wb: Workbook, backtest_result: Dict[str, Any]):
        """创建表现分析工作表"""
        ws = wb.create_sheet(title="表现分析")
        
        # 标题
        ws['A1'] = "各持有期表现分析"
        ws.merge_cells('A1:H1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 表头
        headers = ['持有期', '有效交易', '平均收益%', '胜率%', '最大收益%', '最大亏损%', '标准差%', '中位数收益%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, 'subheader')
        
        # 数据
        performance = backtest_result.get('performance', {})
        holding_periods = performance.get('holding_periods', {})
        
        row = 4
        for period, data in holding_periods.items():
            if 'avg_return' in data:
                ws.cell(row=row, column=1, value=period)
                ws.cell(row=row, column=2, value=data.get('valid_trades', 0))
                ws.cell(row=row, column=3, value=self._format_number(data.get('avg_return', 0)))
                ws.cell(row=row, column=4, value=self._format_number(data.get('positive_rate', 0)))
                ws.cell(row=row, column=5, value=self._format_number(data.get('max_return', 0)))
                ws.cell(row=row, column=6, value=self._format_number(data.get('min_return', 0)))
                ws.cell(row=row, column=7, value=self._format_number(data.get('std_return', 0)))
                ws.cell(row=row, column=8, value=self._format_number(data.get('median_return', 0)))
                
                # 应用条件格式
                avg_return_cell = ws.cell(row=row, column=3)
                if data.get('avg_return', 0) > 0:
                    self._apply_cell_style(avg_return_cell, 'positive')
                else:
                    self._apply_cell_style(avg_return_cell, 'negative')
                
                row += 1
        
        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
    
    def _create_selection_samples_sheet(self, wb: Workbook, backtest_result: Dict[str, Any]):
        """创建选股样本工作表"""
        ws = wb.create_sheet(title="选股样本")
        
        # 标题
        ws['A1'] = "选股样本 (部分数据)"
        ws.merge_cells('A1:D1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 表头
        headers = ['股票代码', '选股日期', '选股价格', '策略']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, 'subheader')
        
        # 数据
        selections_sample = backtest_result.get('selections_sample', [])
        
        for row_idx, selection in enumerate(selections_sample, 4):
            ws.cell(row=row_idx, column=1, value=selection.get('symbol', ''))
            ws.cell(row=row_idx, column=2, value=selection.get('selection_date', ''))
            ws.cell(row=row_idx, column=3, value=selection.get('selection_price', 0))
            ws.cell(row=row_idx, column=4, value=selection.get('strategy', ''))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
    
    def create_comprehensive_report(self, strategy_names: List[str], 
                                  include_backtest: bool = True) -> str:
        """
        创建综合报告
        
        Args:
            strategy_names: 策略名称列表
            include_backtest: 是否包含回测
            
        Returns:
            报告文件路径
        """
        try:
            logger.info(f"开始生成综合报告，策略: {strategy_names}")
            
            # 创建Excel工作簿
            wb = Workbook()
            wb.remove(wb.active)
            
            # 1. 创建总览工作表
            self._create_comprehensive_overview_sheet(wb, strategy_names)
            
            # 2. 为每个策略创建选股结果工作表
            for strategy_name in strategy_names:
                try:
                    results = self.strategy_engine.execute_strategy(strategy_name, 30)
                    if not results.empty:
                        self._create_strategy_sheet(wb, strategy_name, results)
                except Exception as e:
                    logger.error(f"处理策略 {strategy_name} 时出错: {e}")
            
            # 3. 如果需要，添加回测结果
            if include_backtest:
                self._create_backtest_comparison_sheet(wb, strategy_names)
            
            # 保存文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{self.output_dir}/comprehensive_report_{timestamp}.xlsx"
            wb.save(filename)
            
            logger.info(f"综合报告已生成: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"生成综合报告失败: {e}")
            return ""
    
    def _create_comprehensive_overview_sheet(self, wb: Workbook, strategy_names: List[str]):
        """创建综合总览工作表"""
        ws = wb.create_sheet(title="总览")
        
        # 标题
        ws['A1'] = "短线选股系统综合报告"
        ws.merge_cells('A1:F1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 基本信息
        ws['A3'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"包含策略: {', '.join(strategy_names)}"
        ws['A5'] = f"策略数量: {len(strategy_names)}"
        
        # 策略列表
        row = 7
        ws.cell(row=row, column=1, value="策略列表").font = Font(bold=True)
        row += 1
        
        for i, strategy_name in enumerate(strategy_names, 1):
            strategies = self.strategy_engine.get_available_strategies()
            strategy_config = strategies.get(strategy_name, {})
            
            ws.cell(row=row, column=1, value=f"{i}.")
            ws.cell(row=row, column=2, value=strategy_name)
            ws.cell(row=row, column=3, value=strategy_config.get('name', ''))
            ws.cell(row=row, column=4, value=strategy_config.get('description', ''))
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
    
    def _create_strategy_sheet(self, wb: Workbook, strategy_name: str, results: pd.DataFrame):
        """为单个策略创建工作表"""
        # 限制工作表名称长度
        sheet_name = strategy_name[:25] if len(strategy_name) > 25 else strategy_name
        ws = wb.create_sheet(title=sheet_name)
        
        # 标题
        ws['A1'] = f"策略: {strategy_name}"
        ws.merge_cells('A1:H1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 表头
        headers = ['排名', '股票代码', '股票名称', '综合评分', '技术评分', '动量评分', '成交量评分', '5日涨幅%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, 'subheader')
        
        # 数据
        for row_idx, (_, row) in enumerate(results.head(20).iterrows(), 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=row.get('symbol', ''))
            ws.cell(row=row_idx, column=3, value=row.get('name', ''))
            ws.cell(row=row_idx, column=4, value=row.get('comprehensive_score', 0))
            ws.cell(row=row_idx, column=5, value=row.get('technical_score', 0))
            ws.cell(row=row_idx, column=6, value=row.get('momentum_score', 0))
            ws.cell(row=row_idx, column=7, value=row.get('volume_score', 0))
            ws.cell(row=row_idx, column=8, value=row.get('price_change_5d', 0))
        
        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
    
    def _create_backtest_comparison_sheet(self, wb: Workbook, strategy_names: List[str]):
        """创建回测比较工作表"""
        ws = wb.create_sheet(title="回测比较")
        
        # 标题
        ws['A1'] = "策略回测比较"
        ws.merge_cells('A1:F1')
        self._apply_cell_style(ws['A1'], 'header')
        
        # 表头
        headers = ['策略名称', '平均收益%', '胜率%', '夏普比率', '策略评级', '风险水平']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, 'subheader')
        
        # 简化的回测数据（实际应用中需要真实回测）
        row = 4
        for strategy_name in strategy_names:
            ws.cell(row=row, column=1, value=strategy_name)
            ws.cell(row=row, column=2, value="待回测")
            ws.cell(row=row, column=3, value="待回测")
            ws.cell(row=row, column=4, value="待回测")
            ws.cell(row=row, column=5, value="待回测")
            ws.cell(row=row, column=6, value="待回测")
            row += 1
        
        # 调整列宽
        for col in range(1, 7):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15


def main():
    """测试增强版输出管理器"""
    from database import DatabaseManager
    
    # 初始化
    db = DatabaseManager()
    output_manager = EnhancedOutputManager(db)
    
    print("=== 增强版输出管理器测试 ===")
    
    # 1. 测试策略选股报告
    print("\n1. 生成策略选股报告...")
    strategy_name = 'momentum_breakout'
    
    try:
        report_file = output_manager.create_strategy_selection_report(strategy_name, max_results=20)
        if report_file:
            print(f"策略选股报告已生成: {report_file}")
        else:
            print("策略选股报告生成失败")
    except Exception as e:
        print(f"生成策略选股报告时出错: {e}")
    
    # 2. 测试综合报告
    print("\n2. 生成综合报告...")
    strategy_names = ['momentum_breakout', 'technical_reversal']
    
    try:
        comprehensive_file = output_manager.create_comprehensive_report(
            strategy_names,
            include_backtest=False
        )
        if comprehensive_file:
            print(f"综合报告已生成: {comprehensive_file}")
        else:
            print("综合报告生成失败")
    except Exception as e:
        print(f"生成综合报告时出错: {e}")
    
    # 3. 测试回测报告
    print("\n3. 生成回测报告...")
    
    try:
        backtest_file = output_manager.create_backtest_report(
            strategy_name,
            '2024-01-01',
            '2024-03-31'
        )
        if backtest_file:
            print(f"回测报告已生成: {backtest_file}")
        else:
            print("回测报告生成失败")
    except Exception as e:
        print(f"生成回测报告时出错: {e}")


if __name__ == "__main__":
    main()