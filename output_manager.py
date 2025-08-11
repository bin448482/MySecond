"""
输出管理模块
负责选股结果的格式化输出和Excel文件生成
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Dict, List, Optional, Any
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from database import DatabaseManager
from utils import config_manager, format_number, format_percentage, get_date_string

logger = logging.getLogger(__name__)


class OutputManager:
    """输出管理器类"""
    
    def __init__(self, db_manager: DatabaseManager):
        """
        初始化输出管理器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db = db_manager
        self.config = config_manager
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """确保输出目录存在"""
        output_dir = self.config.get('output.output_dir', 'output')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(f"创建输出目录: {output_dir}")
    
    def format_selection_results(self, results: pd.DataFrame) -> pd.DataFrame:
        """
        格式化选股结果
        
        Args:
            results: 原始选股结果
            
        Returns:
            格式化后的结果DataFrame
        """
        if results.empty:
            return results
        
        formatted_results = results.copy()
        
        # 重新排列和重命名列
        column_mapping = {
            'symbol': '股票代码',
            'name': '股票名称',
            'current_price': '当前价格',
            'price_change': '涨跌幅(%)',
            'volume_ratio': '量比',
            'turnover_rate': '换手率(%)',
            'macd_signal': 'MACD信号',
            'rsi_signal': 'RSI信号',
            'ma_signal': '均线信号',
            'total_score': '综合得分'
        }
        
        # 选择需要的列
        available_columns = [col for col in column_mapping.keys() if col in formatted_results.columns]
        formatted_results = formatted_results[available_columns]
        
        # 重命名列
        formatted_results.rename(columns=column_mapping, inplace=True)
        
        # 格式化数值
        if '当前价格' in formatted_results.columns:
            formatted_results['当前价格'] = formatted_results['当前价格'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        
        if '涨跌幅(%)' in formatted_results.columns:
            formatted_results['涨跌幅(%)'] = formatted_results['涨跌幅(%)'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        
        if '量比' in formatted_results.columns:
            formatted_results['量比'] = formatted_results['量比'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        
        if '换手率(%)' in formatted_results.columns:
            formatted_results['换手率(%)'] = formatted_results['换手率(%)'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        
        if '综合得分' in formatted_results.columns:
            formatted_results['综合得分'] = formatted_results['综合得分'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        
        return formatted_results
    
    def generate_summary_statistics(self, results: pd.DataFrame) -> Dict[str, Any]:
        """
        生成统计摘要
        
        Args:
            results: 选股结果
            
        Returns:
            统计摘要字典
        """
        if results.empty:
            return {}
        
        summary = {
            '选股时间': get_date_string(),
            '选出股票数量': len(results),
            '平均涨幅': results['price_change'].mean() if 'price_change' in results.columns else 0,
            '最大涨幅': results['price_change'].max() if 'price_change' in results.columns else 0,
            '最小涨幅': results['price_change'].min() if 'price_change' in results.columns else 0,
            '平均量比': results['volume_ratio'].mean() if 'volume_ratio' in results.columns else 0,
            '平均换手率': results['turnover_rate'].mean() if 'turnover_rate' in results.columns else 0,
            '平均综合得分': results['total_score'].mean() if 'total_score' in results.columns else 0
        }
        
        # 技术指标分布统计
        if 'macd_signal' in results.columns:
            macd_counts = results['macd_signal'].value_counts()
            summary['MACD信号分布'] = macd_counts.to_dict()
        
        if 'rsi_signal' in results.columns:
            rsi_counts = results['rsi_signal'].value_counts()
            summary['RSI信号分布'] = rsi_counts.to_dict()
        
        if 'ma_signal' in results.columns:
            ma_counts = results['ma_signal'].value_counts()
            summary['均线信号分布'] = ma_counts.to_dict()
        
        return summary
    
    def create_excel_workbook(self, results: pd.DataFrame, 
                            summary: Dict[str, Any]) -> Workbook:
        """
        创建Excel工作簿
        
        Args:
            results: 选股结果
            summary: 统计摘要
            
        Returns:
            Excel工作簿对象
        """
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 获取工作表配置
        sheet_config = self.config.get('output.excel_format.sheets', {})
        
        # 1. 创建选股结果工作表
        self._create_results_sheet(wb, results, sheet_config.get('main', '选股结果'))
        
        # 2. 创建统计摘要工作表
        self._create_summary_sheet(wb, summary, sheet_config.get('summary', '统计摘要'))
        
        # 3. 创建技术指标详情工作表（如果有数据）
        if not results.empty:
            self._create_technical_sheet(wb, results, sheet_config.get('technical', '技术指标'))
        
        return wb
    
    def _create_results_sheet(self, wb: Workbook, results: pd.DataFrame, sheet_name: str):
        """创建选股结果工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        if results.empty:
            ws['A1'] = "暂无选股结果"
            return
        
        # 格式化结果
        formatted_results = self.format_selection_results(results)
        
        # 写入数据
        for r in dataframe_to_rows(formatted_results, index=False, header=True):
            ws.append(r)
        
        # 设置样式
        self._apply_results_styling(ws, len(formatted_results))
    
    def _create_summary_sheet(self, wb: Workbook, summary: Dict[str, Any], sheet_name: str):
        """创建统计摘要工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        row = 1
        
        # 基础统计信息
        ws[f'A{row}'] = "选股统计摘要"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        row += 2
        
        basic_stats = [
            ('选股时间', summary.get('选股时间', 'N/A')),
            ('选出股票数量', summary.get('选出股票数量', 0)),
            ('平均涨幅(%)', f"{summary.get('平均涨幅', 0):.2f}"),
            ('最大涨幅(%)', f"{summary.get('最大涨幅', 0):.2f}"),
            ('最小涨幅(%)', f"{summary.get('最小涨幅', 0):.2f}"),
            ('平均量比', f"{summary.get('平均量比', 0):.2f}"),
            ('平均换手率(%)', f"{summary.get('平均换手率', 0):.2f}"),
            ('平均综合得分', f"{summary.get('平均综合得分', 0):.2f}")
        ]
        
        for stat_name, stat_value in basic_stats:
            ws[f'A{row}'] = stat_name
            ws[f'B{row}'] = stat_value
            row += 1
        
        row += 1
        
        # 技术指标分布
        signal_distributions = [
            ('MACD信号分布', summary.get('MACD信号分布', {})),
            ('RSI信号分布', summary.get('RSI信号分布', {})),
            ('均线信号分布', summary.get('均线信号分布', {}))
        ]
        
        for dist_name, dist_data in signal_distributions:
            if dist_data:
                ws[f'A{row}'] = dist_name
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for signal, count in dist_data.items():
                    ws[f'A{row}'] = f"  {signal}"
                    ws[f'B{row}'] = count
                    row += 1
                
                row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_technical_sheet(self, wb: Workbook, results: pd.DataFrame, sheet_name: str):
        """创建技术指标详情工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 选择技术指标相关列
        tech_columns = ['symbol', 'name', 'macd', 'rsi', 'volume_ratio_tech', 
                       'macd_signal', 'rsi_signal', 'ma_signal']
        
        available_tech_columns = [col for col in tech_columns if col in results.columns]
        
        if not available_tech_columns:
            ws['A1'] = "暂无技术指标数据"
            return
        
        tech_data = results[available_tech_columns].copy()
        
        # 重命名列
        tech_column_mapping = {
            'symbol': '股票代码',
            'name': '股票名称',
            'macd': 'MACD值',
            'rsi': 'RSI值',
            'volume_ratio_tech': '量比',
            'macd_signal': 'MACD信号',
            'rsi_signal': 'RSI信号',
            'ma_signal': '均线信号'
        }
        
        tech_data.rename(columns=tech_column_mapping, inplace=True)
        
        # 格式化数值
        if 'MACD值' in tech_data.columns:
            tech_data['MACD值'] = tech_data['MACD值'].apply(lambda x: f"{x:.4f}" if pd.notna(x) else "N/A")
        
        if 'RSI值' in tech_data.columns:
            tech_data['RSI值'] = tech_data['RSI值'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        
        # 写入数据
        for r in dataframe_to_rows(tech_data, index=False, header=True):
            ws.append(r)
        
        # 设置样式
        self._apply_technical_styling(ws, len(tech_data))
    
    def _apply_results_styling(self, ws, data_rows: int):
        """应用选股结果样式"""
        # 标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 数据行样式
        for row in range(2, data_rows + 2):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 交替行颜色
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 设置列宽
        column_widths = self.config.get('output.excel_format.column_widths', {})
        default_width = 12
        
        for col_num, column in enumerate(ws.columns, 1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            header_value = ws.cell(row=1, column=col_num).value
            
            # 根据列名设置宽度
            if header_value == '股票代码':
                ws.column_dimensions[column_letter].width = column_widths.get('symbol', 10)
            elif header_value == '股票名称':
                ws.column_dimensions[column_letter].width = column_widths.get('name', 15)
            elif '价格' in str(header_value):
                ws.column_dimensions[column_letter].width = column_widths.get('price', 12)
            elif '得分' in str(header_value):
                ws.column_dimensions[column_letter].width = column_widths.get('score', 10)
            else:
                ws.column_dimensions[column_letter].width = default_width
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=data_rows + 1):
            for cell in row:
                cell.border = thin_border
    
    def _apply_technical_styling(self, ws, data_rows: int):
        """应用技术指标样式"""
        # 复用结果样式
        self._apply_results_styling(ws, data_rows)
        
        # 特殊处理RSI值的颜色
        rsi_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'RSI值':
                rsi_col = col
                break
        
        if rsi_col:
            for row in range(2, data_rows + 2):
                cell = ws.cell(row=row, column=rsi_col)
                try:
                    rsi_value = float(cell.value)
                    if rsi_value > 70:
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 红色-超买
                    elif rsi_value < 30:
                        cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")  # 绿色-超卖
                except (ValueError, TypeError):
                    pass
    
    def export_to_excel(self, results: pd.DataFrame, filename: Optional[str] = None) -> str:
        """
        导出选股结果到Excel文件
        
        Args:
            results: 选股结果DataFrame
            filename: 文件名，如果为None则自动生成
            
        Returns:
            导出的文件路径
        """
        if filename is None:
            filename_template = self.config.get('output.excel_filename', 'selected_stocks_{date}.xlsx')
            filename = filename_template.format(date=get_date_string(format_str='%Y%m%d'))
        
        output_dir = self.config.get('output.output_dir', 'output')
        filepath = os.path.join(output_dir, filename)
        
        try:
            # 生成统计摘要
            summary = self.generate_summary_statistics(results)
            
            # 创建Excel工作簿
            wb = self.create_excel_workbook(results, summary)
            
            # 保存文件
            wb.save(filepath)
            
            logger.info(f"选股结果已导出到: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"导出Excel文件失败: {e}")
            raise
    
    def save_results_to_database(self, results: pd.DataFrame) -> int:
        """
        保存选股结果到数据库
        
        Args:
            results: 选股结果DataFrame
            
        Returns:
            保存的记录数
        """
        try:
            return self.db.save_selection_results(results)
        except Exception as e:
            logger.error(f"保存选股结果到数据库失败: {e}")
            return 0
    
    def generate_performance_report(self, days: int = 30) -> Dict[str, Any]:
        """
        生成历史表现报告
        
        Args:
            days: 统计天数
            
        Returns:
            表现报告字典
        """
        try:
            conn = self.db.get_connection()
            
            # 获取历史选股结果
            query = '''
                SELECT * FROM selection_results 
                WHERE date >= date('now', '-{} days')
                ORDER BY date DESC, total_score DESC
            '''.format(days)
            
            hist_results = pd.read_sql_query(query, conn)
            conn.close()
            
            if hist_results.empty:
                return {'message': f'过去{days}天没有选股记录'}
            
            # 统计分析
            report = {
                '统计期间': f'过去{days}天',
                '总选股次数': len(hist_results['date'].unique()),
                '总选股数量': len(hist_results),
                '平均每次选股数量': len(hist_results) / len(hist_results['date'].unique()),
                '平均涨幅': hist_results['price_change'].mean(),
                '平均量比': hist_results['volume_ratio'].mean(),
                '平均换手率': hist_results['turnover_rate'].mean(),
                '平均综合得分': hist_results['total_score'].mean()
            }
            
            # 按日期统计
            daily_stats = hist_results.groupby('date').agg({
                'symbol': 'count',
                'price_change': 'mean',
                'total_score': 'mean'
            }).rename(columns={'symbol': 'count'})
            
            report['每日统计'] = daily_stats.to_dict('index')
            
            return report
            
        except Exception as e:
            logger.error(f"生成表现报告失败: {e}")
            return {'error': str(e)}
    
    def print_results_summary(self, results: pd.DataFrame):
        """
        打印选股结果摘要到控制台
        
        Args:
            results: 选股结果DataFrame
        """
        if results.empty:
            print("没有找到符合条件的股票")
            return
        
        print(f"\n{'='*60}")
        print(f"选股结果摘要 - {get_date_string()}")
        print(f"{'='*60}")
        
        print(f"共选出 {len(results)} 只股票")
        
        if 'price_change' in results.columns:
            print(f"平均涨幅: {results['price_change'].mean():.2f}%")
            print(f"涨幅范围: {results['price_change'].min():.2f}% ~ {results['price_change'].max():.2f}%")
        
        if 'total_score' in results.columns:
            print(f"平均得分: {results['total_score'].mean():.2f}")
            print(f"得分范围: {results['total_score'].min():.2f} ~ {results['total_score'].max():.2f}")
        
        print(f"\n前5名股票:")
        print("-" * 60)
        
        for i, (_, row) in enumerate(results.head(5).iterrows(), 1):
            symbol = row.get('symbol', 'N/A')
            name = row.get('name', 'N/A')
            price_change = row.get('price_change', 0)
            score = row.get('total_score', 0)
            
            print(f"{i}. {symbol} {name}")
            print(f"   涨幅: {price_change:.2f}% | 得分: {score:.2f}")
        
        print(f"{'='*60}\n")


def main():
    """测试输出管理功能"""
    from database import DatabaseManager
    from stock_selector import StockSelector
    
    # 初始化
    db = DatabaseManager()
    selector = StockSelector(db)
    output_manager = OutputManager(db)
    
    print("测试输出管理功能...")
    
    # 执行选股（获取少量结果用于测试）
    results = selector.select_stocks(max_results=5)
    
    if not results.empty:
        # 打印摘要
        output_manager.print_results_summary(results)
        
        # 导出Excel
        try:
            filepath = output_manager.export_to_excel(results)
            print(f"Excel文件已导出: {filepath}")
        except Exception as e:
            print(f"Excel导出失败: {e}")
        
        # 保存到数据库
        saved_count = output_manager.save_results_to_database(results)
        print(f"已保存 {saved_count} 条记录到数据库")
        
    else:
        print("没有选股结果可供导出")


if __name__ == "__main__":
    main()